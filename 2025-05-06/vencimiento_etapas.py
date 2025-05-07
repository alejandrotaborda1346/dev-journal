import pandas as pd
import os
from datetime import datetime


script_dir = os.path.dirname(os.path.abspath(__file__))
path_excel = os.path.join(script_dir, "..", "Documentos_excel", "Extracción_GQE.xlsx")

df = pd.read_excel(path_excel,
                   sheet_name =  "GQE",
                   header     =  0,)

# Crear columna combinada como en Excel, ='[Extracción GQE.xlsx]GQE'!$A2&"  ["&'[Extracción GQE.xlsx]GQE'!$FS2&"]"
# Recuerda que, .astype(str) → por si algún valor es numérico o NaN
#df['REQ_con_prioridad'] = df.iloc[:, 0].astype(str) + "  [" + df.iloc[:, 175].astype(str) + "]"
#df['REQ_con_prioridad'] = df.iloc[:, 0].fillna('').astype(str) + "  [" + df.iloc[:, 175].fillna('').astype(str) + "]"


#Columna NC Number
df["NC Number"]         = df.iloc[:,0].fillna('').astype(str) + "  [" + df['Ranking'].fillna('').astype(str) + "]"

#Columna Age
df["Age"]               = df.iloc[:,2]

#Columna Status
df["Status"]            = df.iloc[:,3]

#Columna Additional status
df["Additional status"] = df.iloc[:,4]

#Columna creator
df["Creator"]           = df.iloc[:,6]

#Columna Supplier name
df["Part number"]       = df.iloc[:,18]


#Columna Supplier name
df["Part name"]         = df.iloc[:,19]



#Columna Supplier name
df["Supplier name"]     = df.iloc[:,32]




#Columna Creation date
df["Creation date"]    = df.iloc[:,7]
#df["Creation date"] = pd.to_datetime(df["Creation date"], errors="coerce")





#Columna Steps 1-4 sending deadline
# Evitar errores si hay NaN usando fillna('')
df["Steps 1-4 sending deadline"] = df.apply(    #Usa df.apply(..., axis=1) para recorrer fila por fila.
    lambda row: row.iloc[79] if (               #Si todas las condiciones se cumplen, devuelve el valor de la columna CB (iloc[79]).
        row.iloc[3] != "Abandonada" and         #lambda row:Es una función anónima en Python, usada para crear funciones rápidas y pequeñas sin necesidad de usar def
        row.iloc[3] != "Cerrada" and            #row representa una fila del DataFrame.
        row.iloc[4] != "Suspendida"             #lambda row: ... define una función que recibe esa fila y devuelve algo
    ) else "",
    axis=1
)
df["Steps 1-4 sending deadline"] = pd.to_datetime(df["Steps 1-4 sending deadline"], errors="coerce")



#columna Steps 5-6 sending deadline
df["Steps 5-6 sending deadline"] = df.apply(
    lambda row: row.iloc[80] if (
        row.iloc[3] != "Abandonada" and
        row.iloc[3] != "Cerrada"    and
        row.iloc[4] != "Suspendida"
    ) else "",
    axis = 1
)
# Convertir a datetime
df["Steps 5-6 sending deadline"] = pd.to_datetime(df["Steps 5-6 sending deadline"], errors="coerce")
#pd.to_datetime, covierte str o valores numericos a fechas, esto puede servir para comparar fechas despues
#errors="coerce", convierte valores no válidos (como cadenas vacías, texto, etc.) en NaT (Not a Time), evitando errores.





#columna Steps 5-6 sending date
df["Steps 5-6 sending date"] = df["Steps 5-6 sending date"].astype(str).fillna("")
df["Steps 5-6 sending date"] = pd.to_datetime(df["Steps 5-6 sending date"], errors="coerce")








#columna Steps 7-8 sending deadline
#=SI.ERROR(SI('[Extracción GQE.xlsx]GQE'!$D2<>"Abandonada";SI('[Extracción GQE.xlsx]GQE'!$D2<>"Cerrada";SI('[Extracción GQE.xlsx]GQE'!$E2<>"Suspendida";('[Extracción GQE.xlsx]GQE'!$CD2);"");"");"");"")
df["Steps 7-8 sending deadline"] = df.apply(
    lambda row: row.iloc[81] if (
        row.iloc[3] != "Abandonada" and
        row.iloc[3] != "Cerrada"    and
        row.iloc[4] != "Suspendida"
    ) else "",
    axis = 1
)   
# Convertir a datetime
df["Steps 7-8 sending deadline"] = pd.to_datetime(df["Steps 7-8 sending deadline"], errors="coerce")




#Columna Steps 7-8 sending date
df["Steps 7-8 sending date"] = df["Steps 7-8 sending date"].astype(str).fillna("")
df["Steps 7-8 sending date"] = pd.to_datetime(df["Steps 7-8 sending date"], errors="coerce")





#VARIABLE PARA CALCULAR SEMANA ACTUAL
from datetime import date
semana_actual = 18 #date.today().isocalendar()[1]

#Columna Semana límite 1-4  =SI(E2="";"";NUM.DE.SEMANA(E2;21))
df["Semana límite 1-4"] = df["Steps 1-4 sending deadline"].apply(
    lambda x: x.isocalendar().week if pd.notnull(x) else ""
)


#Columna Semana límite 5-6 
df["Semana límite 5-6"] = df["Steps 5-6 sending deadline"].apply(
    lambda x: x.isocalendar().week if pd.notnull(x) else ""
)


##Columna Semana límite 7-8 
df["Semana límite 7-8"] = df["Steps 7-8 sending deadline"].apply(
    lambda x: x.isocalendar().week if pd.notnull(x) else ""
)





#Columna Technical agreement date
df["Technical agreement date"] = df.apply(
    lambda row: row.iloc[10] if (
        row.iloc[3] != "Abandonada" and
        row.iloc[3] != "Cerrada"    
    ) else "",
    axis = 1
) 
#df["Technical agreement date"] = pd.to_datetime(df["Technical agreement date"], errors="coerce")
df["Technical agreement date"] = df["Technical agreement date"].astype(str).fillna("")
df["Technical agreement date"] = pd.to_datetime(df["Technical agreement date"], errors="coerce")




#Columna IF Status

df["IF Status"] = df["IF Amount Agreement Date"].astype(str).fillna("")
df["IF Status"] = pd.to_datetime(df["IF Status"], errors="coerce")



#Columna STATUS DF

df["STATUS DF"] = df.apply(lambda row: (
    # 1. Si tipo = POI → "NA"
    "NA" if row.iloc[36] == "POI" else

    # 2. Si Q2 (Cancellation date) tiene valor, R2 (Cancellation reason) está vacío y HOY > Q2 → "PTE Abrir DF"
    "PTE Abrir DF" if (
        pd.notna(row["Cancellation date"]) and          #pd.notna(valor) → devuelve True si el valor no está vacío (no es NaN).
        row["Cancellation reason"] == "" and            #pd.isna(valor) → devuelve True si el valor está vacío (es NaN)
        datetime.today() > row["Cancellation date"]
    ) else

    # 3. Si tipo = POE, DL vacío, DM distinto de "Abandonada", Q2 tiene fecha, R2 = "Acuerdo Principio", y vencido → "PTE Enviar DF"
    "PTE Enviar DF" if (
        row.iloc[36]  == "POE" and                  # AK2
        row.iloc[116] == "" and                    # DL2
        row.iloc[117] != "Abandonada" and          # DM2
        pd.notna(row["Cancellation date"]) and              # Q2
        row["Cancellation reason"] == "Acuerdo Principio" and   # R2
        datetime.today() > row["Cancellation datene"]
    ) else

    # 4. Si Q2 y R2 tienen valor → "OK"
    "OK" if (
        pd.notna(row["Cancellation date"]) and row["Cancellation reason"] != ""
    ) else

    # 5. Ninguna se cumple → ""
    ""
), axis=1)





#Columna 8D sending & confirmation status
df["8D sending & confirmation status"] = df.apply(
    lambda row: (
        str(row.iloc[90])[:2] + "-" + str(row.iloc[91])[:2]
        if row.iloc[3] != "Abandonada" and
           row.iloc[3] != "Cerrada" and
           row.iloc[4] != "Suspendida"
        else ""
    ),
    axis=1
)





#Columna IF Status 2
df["IF Status 2"] = df.apply(
    lambda row: "AB" if row.iloc[116] == "AB" else row.iloc[117],
    axis=1
)


#columna STATUS DF 2
df["STATUS DF 2"] = df.apply(
    lambda row: (
        "NA" if (
            row.iloc[3] == "Abandonada" or
            row.iloc[21] == "AB" or
            row.iloc[18] == "NA"
        ) else
        "OK" if (
            row.iloc[21] in ["AM", "FM", "SD"]
        ) else
        "P" if (
            row.iloc[18] == "OK" and
            row.iloc[21] == "AP" and
            row.iloc[20] == "E7-V7"
        ) else
        ""
    ),
    axis=1
)



#columna RECHAZO 1-4
df["RECHAZO 1-4"]  = df.iloc[:,83]
df["RECHAZO 1-4"]  = pd.to_datetime(df["RECHAZO 1-4"], errors="coerce")


#columna RECHAZO 5-6
df["RECHAZO 5-6"]  = df.iloc[:,84]
df["RECHAZO 5-6"]  = pd.to_datetime(df["RECHAZO 5-6"], errors="coerce")


#columna RECHAZO 7-8
df["RECHAZO 7-8"]  = df.iloc[:,85]
df["RECHAZO 7-8"]  = pd.to_datetime(df["RECHAZO 7-8"], errors="coerce")


#columna POE/POI
df["POE/POI"]      = df.iloc[:,36]


#columna 1-4
df["1-4"]          = df.iloc[:,92].fillna(0).astype(int)


#columna 1-4
df["5-6"]          = df.iloc[:,93].fillna(0).astype(int)


#columna 1-4
df["7-8"]          = df.iloc[:,94].fillna(0).astype(int)




#Filtros para el numero de la semana en la que estamos


df_1_4 = df[df["Semana límite 1-4"] == semana_actual][
    ["NC Number", "Creator", "Supplier name", "Steps 1-4 sending deadline", "1-4"]
]

df_5_6 = df[df["Semana límite 5-6"] == semana_actual][
    ["NC Number", "Creator", "Supplier name", "Steps 5-6 sending deadline", "5-6", "RECHAZO 5-6"]
]

df_7_8 = df[df["Semana límite 7-8"] == semana_actual][
    ["NC Number", "Creator", "Supplier name", "Steps 7-8 sending deadline", "7-8", "RECHAZO 7-8"]
]




"""print(
    df[["NC Number", "Supplier name", "Creator", "Creation date","Steps 1-4 sending deadline","Steps 1-4 sending date","Steps 5-6 sending deadline",
        "Steps 5-6 sending date", "Steps 7-8 sending deadline" ,"Steps 7-8 sending date", "Technical agreement date","IF Status", "STATUS DF",
        "8D sending & confirmation status", "IF Status 2", "STATUS DF 2", "RECHAZO 1-4", "RECHAZO 5-6", "RECHAZO 7-8", "1-4", "5-6", "7-8"]].head(10)
)"""




# print(df[["Semana límite 7-8", "Semana límite 5-6", "Semana límite 7-8"]].head(5))
#print( df["Steps 1-4 sending deadline"].head(10))



from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import column_index_from_string
import locale #Establece el locale en español para obtener el nombre del día
import datetime




condicion_estado = ~df.iloc[:, 3].isin(["Cerrada", "Abandonada", "DRAFT"])

# Paso 2: Filtrar las que contienen "ANA" en la columna AP
condicion_ana    =  df.iloc[:, 41].astype(str).str.contains("ANA", case=False, na=False)

# Aplicar ambas condiciones
df_filtrado = df[condicion_estado & condicion_ana]

# Ordenar por la columna 3 (índice 2, ya que en pandas se empieza desde 0)
df_ordenado = df_filtrado.sort_values(by=df_filtrado.columns[2])


columnas_deseadas = df_ordenado.iloc[:, [0, 2, 3,4,6,18,19,32]].head(12)


#print(columnas_deseadas.to_string(index=False))

df_SQUALL_ANA = columnas_deseadas.rename(columns={
    "NC Number" : "NC Number",
    "Age" : "Age",
    "Status" : "Status",
    "Additional status" : "Additional status",
    "Creator" : "Creator",
    "Part number" : "Part number",
    "Part name" : "Part name",
    "Supplier name" : "SUPPLIER NAME"
})
df_SQUALL_ANA = df_SQUALL_ANA.head(12)



# Renombrar columnas para visualización
df_VENCIMIENTO_ETAPAS_1_4 = df_1_4.rename(columns={
    "NC Number": "SQUALL",
    "Creator": "CREADOR",
    "Supplier name": "SUPPLIER NAME",
    "Steps 1-4 sending deadline": "FECHA",
    "1-4": "ESTADO"
})

df_VENCIMIENTO_ETAPAS_5_6 = df_5_6.rename(columns={
    "NC Number": "SQUALL",
    "Creator": "CREADOR",
    "Supplier name": "SUPPLIER NAME",
    "Steps 5-6 sending deadline": "FECHA",
    "5-6": "ESTADO"
})

df_VENCIMIENTO_ETAPAS_7_8 = df_7_8.rename(columns={
    "NC Number": "SQUALL",
    "Creator": "CREADOR",
    "Supplier name": "SUPPLIER NAME",
    "Steps 7-8 sending deadline": "FECHA",
    "7-8": "ESTADO"
})



try:
        locale.setlocale(locale.LC_TIME, 'Spanish_Spain')  # Windows
except:
    pass  # Si falla, se usará en inglés

# Función para formatear fecha
def formatear_fecha(fecha):
    if pd.isna(fecha):
        return ""
    dia = fecha.strftime("%a")[:3].lower()  # lun, mar, mié, etc.
    hora = fecha.strftime("%H:%M")          # 10:28
    semana = f"w{fecha.isocalendar()[1]}"   # semana del año
    return f"{dia},({hora}){semana}"

# Aplicar formato personalizado a la columna FECHA
df_VENCIMIENTO_ETAPAS_1_4["FECHA"] = df_VENCIMIENTO_ETAPAS_1_4["FECHA"].apply(formatear_fecha)
df_VENCIMIENTO_ETAPAS_5_6["FECHA"] = df_VENCIMIENTO_ETAPAS_5_6["FECHA"].apply(formatear_fecha)
df_VENCIMIENTO_ETAPAS_7_8["FECHA"] = df_VENCIMIENTO_ETAPAS_7_8["FECHA"].apply(formatear_fecha)





# Ruta del script y archivo de salida
script_dir = os.path.dirname(os.path.abspath(__file__))
path_excel = os.path.join(script_dir, "..", "Output", "ejemplo.xlsx")

# Crear archivo Excel real con hoja vacía
with pd.ExcelWriter(path_excel, engine="openpyxl") as writer:
    pd.DataFrame().to_excel(writer, sheet_name="Hoja1", index=False)

# Abrir archivo Excel
wb = load_workbook(path_excel)
ws = wb["Hoja1"]

# Títulos SQUALL ANA
ws["K2"] = "SQUALL ANA"

# Títulos VENCIMIENTO ETAPAS 1-4
ws["B18"] = "SQUALL"
ws["B19"] = "VENCIMIENTO ETAPAS 1-4"
ws["I19"] = "SIN PROXY"
ws["I20"] = "CLIENT:0%"

# Títulos VENCIMIENTO ETAPAS 5-6
ws["k19"] = "VENCIMIENTO ETAPAS 5-6"
ws["R19"] = "GRAL:0%"
ws["R20"] = "POE:0%"

# Títulos VENCIMIENTO ETAPAS 7-8
ws["T19"] = "VENCIMIENTO ETAPAS 7-8"
ws["AA19"] = "GRAL:0%"
ws["AA20"] = "POE:0%"

# Escribir encabezados en columnas específicas
header_titulos = [
    

    {
        
        "dataframe": df_SQUALL_ANA,
        "headers_row": 3,
        "start_row":   4,
        "column_map": {"NC Number": "K", "Age": "M", "Status": "N", "Additional status": "P", "Creator": "Q", "Part number": "T", "Part name": "V", "SUPPLIER NAME": "Y"}
    },
    {
        
        "dataframe": df_VENCIMIENTO_ETAPAS_1_4,
        "headers_row": 21,
        "start_row": 22,
        "column_map": {"SQUALL": "B", "CREADOR": "D", "SUPPLIER NAME": "F", "FECHA": "H", "ESTADO": "I"}
    },
    {
        "dataframe": df_VENCIMIENTO_ETAPAS_5_6,
        "headers_row": 21,
        "start_row": 22,
        "column_map": {"SQUALL": "K", "CREADOR": "M", "SUPPLIER NAME": "O", "FECHA": "Q", "ESTADO": "R"}
    },
    {
        "dataframe": df_VENCIMIENTO_ETAPAS_7_8,
        "headers_row": 21,
        "start_row": 22,
        "column_map": {"SQUALL": "T", "CREADOR": "V", "SUPPLIER NAME": "X", "FECHA": "Z", "ESTADO": "AA"}
    }
]

# Escribir encabezados y datos
for block in header_titulos:
    df_block = block["dataframe"]
    headers_row = block["headers_row"]
    start_row = block["start_row"]
    col_map = block["column_map"]

    for col_name, col_letter in col_map.items():
        col_idx = column_index_from_string(col_letter)
        ws.cell(row=headers_row, column=col_idx, value=col_name)

    for r_idx, row in enumerate(dataframe_to_rows(df_block, index=False, header=False), start=start_row):
        for col_name, col_letter in col_map.items():
            col_idx = column_index_from_string(col_letter)
            ws.cell(row=r_idx, column=col_idx, value=row[list(df_block.columns).index(col_name)])

# Guardar cambios
wb.save(path_excel)





